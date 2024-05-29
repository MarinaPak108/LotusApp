import os
from flaskwebgui import FlaskUI
from datetime import date, datetime
from flask import Flask, redirect, request, render_template
import openpyxl

from models.doctor import Doctor
from models.patient import Patient
from models.record import Record
from service.service import Service

app = Flask(__name__, static_url_path="", static_folder=os.path.join(os.getcwd(), "flask_desktop_app/phrases"))

directory = os.getcwd()

@app.route("/")
def home():
    day = str(date.today())
    wb = Service.getWB()
    names = wb.sheetnames
    if (day not in names):
        return redirect('/assign')
    else:
        ws = wb["current"]
        header = [cell.value for cell in ws[1]]
        records = Service.fromExcelToList(Record, ws)
        Service.countPatients(wb, records, day)
        return render_template('start.html', title="Список отчетов", tab = header, records = records)

@app.route('/assign')
def my_form():
    wb = Service.getWB()
    ws = wb['settings']
    doctors = Service.fromExcelToList(Doctor, ws)
    return render_template('assign.html', doctors = doctors)

@app.route('/assign', methods=['POST'])
def my_form_post():
    docId = request.form['text']
    day = str(date.today())

    wb = Service.getWB()
    
    ws_new = wb.create_sheet(day) 
    ws_header = ["Время","ФИО пациента", "М\Ж\Р", "Дата рождения", "Причина"]
    ws_new.append(ws_header)
    
    wsDoc = wb["settings"]
    docName = wsDoc[docId][1].value
    
    new_data = [day, docName, -1, docId]
    ws = wb["current"]
    ws.append(new_data)
    wb.save("records/medical.xlsx")
    return redirect("/")

@app.route('/day/<id>')
def patients(id):
    wb = Service.getWB()
    isActiveDay = False
    if(id == str(date.today())):
        isActiveDay = True
    if(id in wb.sheetnames):
         ws = wb[id]
         
         patients= Service.fromExcelToList(Patient, ws)
         return render_template('patient.html', patients = patients, day = id, isActive = isActiveDay)
    else:
        return redirect("/")
    
@app.route('/day/<id>', methods=['POST'])
def patients_post(id):
    patient_name = request.form['name']
    patient_type = request.form['type']
    patient_time = datetime.now()
    patient_birthdate = request.form['birthdate']
    patient_reason = request.form['reason']
    patient_pressure = request.form['pressure']
    wb = Service.getWB()
    
    if(id in wb.sheetnames):
        new_data = [patient_time, patient_name, patient_type, patient_birthdate, patient_reason, patient_pressure]
        Service.saveRecord(wb, id, new_data)
        return redirect("/day/"+id)
    else:
        return redirect("/")



if __name__ == "__main__":
  FlaskUI(app=app, server="flask",  width= 800, height=600).run()
  