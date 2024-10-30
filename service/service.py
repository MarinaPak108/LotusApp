import os
import openpyxl
from datetime import date, datetime
from typing import TypeVar
import pandas as pd
import fnmatch
import re
from dateutil.relativedelta import relativedelta

from model.patient import Patient
from model.record import Record
from model.doctor import Doctor
from model.report import Report 
from exception.custom_exception import CustomException

from error.error import Error_msg


class Service():
    T = TypeVar('T')   
    
    def getWB(path_to_file):
        try:
            return openpyxl.load_workbook(path_to_file)
        except Exception as e:
            return "error in getWB:"+str(e)

    def getDay():
        return str(date.today())

    def getTimestamp():
        return datetime.now()
    
    def fromExcelToList(T, ws):
        try: 
            listObjects=[]
            for i in range(2,ws.max_row+1):
                args =[cell.value for cell in ws[i]]
                typeObject = T(*args)
                listObjects.append(typeObject)
            return listObjects
        except Exception as e:
            return "error in fromExcelToList:"+str(e)
        
    def fromExcelToModel(T, ws):
        try:
            args =[cell.value for cell in ws]
            typeObject = T(*args)
            return typeObject
        except Exception as e:
            return "error in fromExcelToModel:"+str(e)
            
    def saveRecord(wb,name, new_data, file_path):
        ws=wb[name]
        ws.append(new_data)
        wb.save(file_path)
     
    def countPatients(self, wb, today, folder_path):
        try: 
            ws = wb['current']
            records = self.fromExcelToList(Record,ws)
        
            for i in  range(2, ws.max_row+1):
                if(records[i-2].patients == -1 or records[i-2].date == today):
                    day_name = records[i-2].date #name of the day and worksheet name
                    men_num=0
                    woman_num=0
                    child_num=0
                    ws_current = wb[day_name] #get to specific day page
                    patients_num = ws_current.max_row-1 #get number of patients that day
                
                    records[i-2].patients = patients_num
                    #if there are patients, then count types
                    if(patients_num!=0):
                        df=pd.read_excel(folder_path, i)#count men woman child
                        mylist = df['М\Ж\Р'].tolist() #get types list
                        #count each type specifically
                        woman_num = mylist.count("Ж") 
                        men_num = mylist.count("М")  
                        child_num = mylist.count("Р")  
                        #save to records list
                        records[i-2].child = child_num
                        records[i-2].woman = woman_num
                        records[i-2].men = men_num
                        #save all info           
                        ws.cell(row=i, column=2).value = patients_num
                        ws.cell(row=i, column=3).value = child_num
                        ws.cell(row=i, column=4).value = woman_num
                        ws.cell(row=i, column=5).value = men_num
        except Exception as e:
            return "error in countPatients:"+str(e)
       
        wb.save(folder_path) #update excel file
        return records
   
    def isAlredySaved(id, name, birthday, file):
        try:
            #to get list of saved names and birthdays
            df=pd.read_excel(file, id) 
            name_list = df["ФИО пациента"].tolist() 
        
            if(name in name_list): # check if name already exists
                return True
            else:
                return False  
        except Exception as e:
            return "error in isAlredySaved:"+str(e) 
        
    def checkSavePatientGetPage(self,
                                id,
                                patient_id, 
                                patient_name, 
                                patient_type, 
                                patient_birthdate, 
                                patient_reason,
                                patient_pressure,
                                patient_doc,
                                patient_docId,
                                file,
                                folder):
        try:
            patient_time = datetime.now()
            wb = self.getWB(file)
            isNotUnique = self.isAlredySaved(id, patient_name, patient_birthdate, file)
            if(isNotUnique):
                return ("/day/"+id+"/1/"+patient_name)
            elif (id in wb.sheetnames):
                new_data = [patient_id, patient_time, patient_name,patient_doc.name, int(patient_docId), patient_type, patient_birthdate, patient_reason, patient_pressure]
                self.formFileSave(self, patient_id, patient_name, patient_pressure, patient_birthdate, patient_reason, patient_doc, folder)
                self.saveRecord(wb, id, new_data, file)
                return ("/day/"+id)
            else:
                return ("/")  
        except Exception as e:
            return "error in checkSavePatientGetPage:"+str(e)  
    
    def formFileSave(self, patient_id, patient_name, patient_pressure, patient_bday, patient_reason, doctor, folder):
        try:
            template = "records/card.xlsx"
            wb= self.getWB(template)
            ws = wb['form']
            #fill doc related fields:
            ##doc name
            ws.cell(row=4, column=7).value=doctor.name
            ##service date
            ws.cell(row=5, column=7).value= '{:%d-%b-%Y}'.format(date.today())
            ##doc spec
            ws.cell(row=6, column=7).value= doctor.spec
            ##nurse
            ws.cell(row=7, column=7).value= doctor.nurse
            
            #fill patient related data:
            ## nomer talona
            ws.cell(row=11, column=3).value= patient_id
            ## patient name
            ws.cell(row=12, column=3).value= patient_name
            ## patient bday
            ws.cell(row=13, column=3).value= patient_bday
            ## patient pressure
            ws.cell(row=14, column=3).value= patient_pressure    
            ## patient reason
            ws.cell(row=16, column=3).value= patient_reason       
            
            p_name = patient_name.replace(" ", "_")
            ##form path to folder:
            doc = os.path.join(folder,patient_id+'.'+doctor.name+'_('+p_name+').xlsx')
            wb.save(doc)
            ## open file
            #os.startfile(doc)
            ##print file
            os.startfile(doc, "print")
        except Exception as e:
            return "error in formFileSave:"+str(e) 
            
    def countGrownUp():
        yrs = date.today()  - relativedelta(years=18)  
        return str(yrs)  
    
    def countCentury():
        yrs = date.today()  - relativedelta(years=100)  
        return str(yrs)          
              
    def getDocName(self, id, file):
        try:
            wb = self.getWB(file)    
            wsDoc = wb["settings"]
            docName = wsDoc[id][1].value
            docSpec = wsDoc[id][2].value
            docNurse = wsDoc[id][3].value
            doc = Doctor(id=id, name=docName, spec=docSpec, nurse= docNurse)
            return doc
        except Exception as e:
            return "error in getDocName:"+str(e) 

    def sortDoctors (docs, patient):
        doc_list = []
        try:
            d = type(docs)
            #get only active doctors by active status
            docs = [d for d in docs if d.active == 1]
            for doc in docs:
                id = doc.id
                sorted_docs=[p for p in patient if p.docId == id]
                doc.num = len(sorted_docs)
                if(doc.num > 0):
                    doc_list.append(doc)
            return docs, doc_list
        except Exception as e:
            return "error in sortDoctors:"+str(e) 
        
    def printFile(folder,patient_id, doctor_name, p_name ):
        doc = os.path.join(folder,patient_id+'.'+doctor_name+'_('+p_name+').xlsx')
        ##print file
        #os.startfile(doc, "print")
        ## open file
        os.startfile(doc)
            
    def countDoctors(self, today, day, folder, report_folder):
        try:
            wb_report = self.getWB(report_folder)
            ws_report = wb_report[day] #get all doctors names in report
            if (day == today):
                df=pd.read_excel(folder, sheet_name=day)
                for i in range(2,ws_report.max_row+1):
                    all_num = 0
                    w_num = 0
                    m_num = 0
                    c_num = 0
                    doc_id = ws_report.cell(i, 1).value
                    #find all records with doctor id:
                    doc_df = df[df['Врач_Индекс'] == doc_id]
                    sz = len(doc_df)
                    if(sz>0):
                        all_num = sz
                        mylist = doc_df['М\Ж\Р'].tolist() #get types list
                        #count each type specifically
                        w_num = mylist.count("Ж") 
                        m_num = mylist.count("М")  
                        c_num = mylist.count("Р") 
                    #save to report file:
                    #save all info           
                    ws_report.cell(row=i, column=5).value = m_num
                    ws_report.cell(row=i, column=6).value = w_num
                    ws_report.cell(row=i, column=7).value = c_num
                    ws_report.cell(row=i, column=8).value = all_num
                wb_report.save(report_folder)  
            reports = self.fromExcelToList(Report, ws_report)
            return reports
        except Exception as e:
            return "error in countDoctors:"+str(e) 
            
    def saveDoctor (self,med_file, report_file, id, name, spec, nurse, isActive):
        day = self.getDay()
        wb = self.getWB(med_file)    
        wsDoc = wb["settings"]
        wbr = self.getWB(report_file)
        wsrDoc = wbr[day]
        name = name.replace(" ", "_")
        #check if records in medical file settings and records file current day are the same:
        if (wsDoc.max_row != wsrDoc.max_row):
            raise CustomException(f"Проверьте записи в файле medical.xls[settings] и файле records.xls[{day}] на соответствие.")
        #check if we need to save new record => get new id
        if(id == ''):
            id = wsDoc.max_row+1
        try:
            wsDoc[id][0].value = id
            wsDoc[id][1].value = name
            wsDoc[id][2].value = spec
            wsDoc[id][3].value = nurse
            wsDoc[id][4].value = isActive
            wb.save(med_file)
            
            wsrDoc[id][0].value = id
            wsrDoc[id][1].value = name
            wsrDoc[id][2].value = spec
            wsrDoc[id][3].value = nurse
            wbr.save(report_file)
        except Exception as e:
             return "error in saveDoctor:"+str(e)
          
    def fromErroToEnum(err_no):
        return 'e_'+str(err_no)
    
    def fromErrorMsgToEnum(err_msg):
        return err_msg.replace(" ", "_")    
            
    def getDoctorList (self, folder, today_folder,  day, docId):
        try: 
            #get doc name
            name = self.getDocName(Service, docId, folder).name
            
            latest_record = 0
            #code to check if already exists, also if exists from what next patient to start
            for file in os.listdir(today_folder):
                if fnmatch.fnmatch(file, "0_%s_*"%(docId)):
                    str_num = re.search("{(.*)}", file)
                    latest_record = int(str_num.group(1))
    
            #in case no records there
            if latest_record == 0:
                df=pd.read_excel(folder, sheet_name=day) 
                latest_record = len(df.axes[0])
            #in case list was already created. adding +1 to start from next record              
            else:
                df=pd.read_excel(folder, sheet_name=day, skiprows = range(1, latest_record+1))                
                latest_record = latest_record + len(df.axes[0])
                
            filtered_patients_list_df = df[df['Врач_Индекс'] == int(docId)].reset_index(drop=True)
            #if latest list was already printed out:
            if df.empty:
                return ("Нет новых данных по врачу c номером id_name: %s_%s" % (docId, name))
            
            msg = self.fillFormPrintPatients(self, 
                                       today_folder, 
                                       docId, 
                                       filtered_patients_list_df,
                                       latest_record,
                                       name)
            return msg                      
        except Exception as e:
            return "error in getDoctorList:"+str(e)
    
    def fillFormPrintPatients(self,
                              folder, 
                              docId, 
                              df_list,
                              last_record,
                              name):
        try:
            template = "records/list.xlsx"
            wb= self.getWB(template)
            ws = wb['list']
            
            ##fill doc related data
            ws.cell(row=1, column=2).value=docId
            ws.cell(row=1, column=4).value = name
            
            ##fill patients related data:
            for index, row in df_list.iterrows():
                index+=3
                #get data from each row
                patient_id = row['ID']
                patient_name = row['ФИО пациента']
                patient_gender = row['М\Ж\Р']
                patient_bday = row['Дата рождения']
                
                #fill created document:
                ws.cell(row = index, column = 1).value=patient_id
                ws.cell(row = index, column = 2).value=patient_name
                ws.cell(row = index, column = 3).value=patient_gender
                ws.cell(row = index, column = 4).value=patient_bday
                
            #save file  
            path = os.path.join(folder,"0_%s_%s_{%s}.xlsx"%(docId,name, last_record))  
            wb.save(path)           
            os.startfile(path, "print") 
            return ("Файл с обновленным списком создан: %s" % path)
        except Exception as e:
            return "error in fillFormPrintPatients:"+str(e)    
                    
            
            
        


